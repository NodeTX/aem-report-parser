import xlrd
from openpyxl import Workbook
from datetime import datetime
import argparse

# data columns
DEVICE_NAME = 1
USER_NAME = 6
SERIAL = 20
BUILD_DATE = 22

parser = argparse.ArgumentParser()
parser.add_argument('-i', help='Input filename (XLS)', required=True)
parser.add_argument('-o', help='Output filename (XLSX)', required=True)
args = parser.parse_args()

input_filename = args.i
output_filename = args.o

if __name__ == '__main__':
    print('Input file: %s | Output file: %s' % (input_filename, output_filename))

    output_wb = Workbook()
    output_sheet = output_wb.active

    input_wb = xlrd.open_workbook(input_filename, formatting_info=True)
    input_sheet_names = input_wb.sheet_names()
    input_sheet = input_wb.sheet_by_name(input_sheet_names[0])  # getting first sheet

    # HEADERS
    output_sheet.cell(row=1, column=1).value = 'Device Name'
    output_sheet.cell(row=1, column=2).value = 'User Name'
    output_sheet.cell(row=1, column=3).value = 'Serial'
    output_sheet.cell(row=1, column=4).value = 'Build Date'

    # COL WIDTHS
    output_sheet.column_dimensions['A'].width = 40
    output_sheet.column_dimensions['B'].width = 45
    output_sheet.column_dimensions['C'].width = 20
    output_sheet.column_dimensions['D'].width = 20

    output_row = 2

    first_dev_name_encountered = False
    num_cols = input_sheet.ncols  # number of columns

    xfx = input_sheet.cell_xf_index(0, 0)
    xf = input_wb.xf_list[xfx]
    EMPTY_CELL_COLOR_CODE = xf.background.pattern_colour_index

    for row_idx in range(0, input_sheet.nrows):
        device_name = input_sheet.cell(row_idx, DEVICE_NAME).value.strip()

        #print(device_name)

        if device_name == '':  # skip empty
            #print('SKIPPING EMPTY DEV NAME')
            continue

        if device_name == 'Device Name':  # find first valid row
            #print('FOUND FIRST DEV NAME')
            first_dev_name_encountered = True

        if not first_dev_name_encountered:
            #print('FIRST DEV NAME NOT ENCOUNTRED')
            continue

        xfx = input_sheet.cell_xf_index(row_idx, DEVICE_NAME)
        xf = input_wb.xf_list[xfx]
        bgx = xf.background.pattern_colour_index

        #print(bgx)
        if bgx != EMPTY_CELL_COLOR_CODE:  # color of 'UNCOLORED' cell
            #print('SKIPP COLORED CELL')
            continue

        pc_name = input_sheet.cell(row_idx, USER_NAME).value.strip()
        serial = input_sheet.cell(row_idx, SERIAL).value.strip()
        build_date = str(input_sheet.cell(row_idx, BUILD_DATE).value).strip()

        if build_date != '' and build_date is not None:
            build_date_fixed_dt = datetime.strptime(build_date, '%d/%m/%y')
        else:
            build_date_fixed_dt = ''

        print('%s | %s | %s | %s' % (device_name, pc_name, serial, build_date))

        output_sheet.cell(row=output_row, column=1).value = device_name
        output_sheet.cell(row=output_row, column=2).value = pc_name
        output_sheet.cell(row=output_row, column=3).value = serial
        output_sheet.cell(row=output_row, column=4).value = build_date_fixed_dt
        output_sheet.cell(row=output_row, column=4).number_format = 'DD/MM/YYYY'

        output_row += 1

    output_wb.save(output_filename)
